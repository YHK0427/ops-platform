"""
벌점/디파짓 리포트 Excel 내보내기 서비스
참조: UnivPT 33기 벌점_디파짓_1주차.xlsx
"""

from __future__ import annotations

import io
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Assignment, Attendance, Ledger, Member, Session, TeamMember

# ── 상수 ──────────────────────────────────────────────────────────────

HEADER_ROW = 8          # 날짜 행
WEEK_ROW = 9            # 주차 행
TITLE_ROW = 10          # 내용 행
DATA_START_ROW = 11     # 멤버 데이터 시작 행
NUM_COL = 3             # C열 = 번호
NAME_COL = 4            # D열 = 이름
DATA_START_COL = 5      # E열 = 첫 번째 세션 데이터

# 스타일
HEADER_FONT = Font(bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14)
HEADER_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
WRAP_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 출석 → 한국어 텍스트 매핑
ATTENDANCE_TEXT: dict[tuple[str, str | None], str] = {
    ("LATE_UNDER10", "PRE"):  "10분 미만 지각\n사전 사유서",
    ("LATE_UNDER10", "POST"): "10분 미만 지각\n사후 사유서O",
    ("LATE_UNDER10", None):   "10분 미만 지각\n사후 사유서X",
    ("LATE_OVER10", "PRE"):   "10분 이상 지각\n사전 사유서",
    ("LATE_OVER10", "POST"):  "10분 이상 지각\n사후 사유서O",
    ("LATE_OVER10", None):    "10분 이상 지각\n사후 사유서X",
    ("EARLY_LEAVE", "PRE"):   "조퇴\n사전 사유서",
    ("EARLY_LEAVE", "POST"):  "조퇴\n사후 사유서O",
    ("EARLY_LEAVE", None):    "조퇴\n사후 사유서X",
    ("ABSENT", "PRE"):        "결석\n사전 사유서",
    ("ABSENT", "POST"):       "결석\n사후 사유서O",
    ("ABSENT", None):         "결석\n사후 사유서X",
    ("EXCUSED", None):        "공결",
    ("EXCUSED", "PRE"):       "공결",
    ("EXCUSED", "POST"):      "공결",
}

# PPT_EMAIL 상태 → 텍스트
PPT_STATUS_TEXT = {
    "LATE": "지각",
    "MISSING": "미제출",
    "EXEMPT": "면제",
}

# 상점 카테고리 (엑셀 컬럼 헤더, description 매칭 패턴)
MERIT_CATEGORIES: list[tuple[str, str]] = [
    ("추상 작성\n(1)", "추억상자 글 작성"),
    ("친바\n(1)", "친바 선정팀"),
    ("리슨업\n(4)", "Listen Up"),
    ("BP\n(4)", "BP"),
    ("피날래 본선진출\n(3)", "피날래 본선 진출"),
    ("발전왕\n(4)", "발전왕 선발"),
    ("오프/오피\n(1)", "오프/오피 선정"),
    ("번개 주최 후 완료\n(2)", "번개 주최 완료"),
    ("리슨업/BP\n베스트\n협력상\n(1)", "베스트 협력상"),
    ("4회 연속\n지각/결석 X\n(2)", "4주 연속 출석 달성"),
    ("번개 참여\n2회\n(1)", "번개 참석"),
    ("정보성\n자료 공유\n(1)", "정보성 자료 공유"),
]

# 출석별 디파짓 차감 금액 (penalty_engine과 동일)
ATTENDANCE_DEPOSIT: dict[tuple[str, str | None], int] = {
    ("LATE_UNDER10", "PRE"): 2000,
    ("LATE_UNDER10", "POST"): 3000,
    ("LATE_UNDER10", None): 4000,
    ("LATE_OVER10", "PRE"): 2000,
    ("LATE_OVER10", "POST"): 3000,
    ("LATE_OVER10", None): 4000,
    ("EARLY_LEAVE", "PRE"): 2000,
    ("EARLY_LEAVE", "POST"): 3000,
    ("EARLY_LEAVE", None): 4000,
    ("ABSENT", "PRE"): 4000,
    ("ABSENT", "POST"): 6000,
    ("ABSENT", None): 8000,
}


# ── 데이터 컨테이너 ──────────────────────────────────────────────────

@dataclass
class ExportData:
    sessions: list[Any]
    members: list[Any]
    att_map: dict[int, dict[int, Any]]                       # member_id → session_id → Attendance
    assign_map: dict[int, dict[int, dict[str, Any]]]         # member_id → session_id → type → Assignment
    team_ppt_map: dict[int, dict[int, str]]                  # session_id → member_id → status (TEAM PPT)
    ledger_by_session: dict[int, dict[int, list[Any]]]       # member_id → session_id → [Ledger]
    ledger_no_session: dict[int, list[Any]] = field(default_factory=dict)


# ── 데이터 수집 ──────────────────────────────────────────────────────

async def gather_export_data(db: AsyncSession) -> ExportData:
    # 1. 세션 (날짜순)
    sessions = (await db.execute(
        select(Session).order_by(Session.date.asc())
    )).scalars().all()

    # 2. 활성 멤버 (이름순)
    members = (await db.execute(
        select(Member).where(Member.is_active == True).order_by(Member.name)  # noqa: E712
    )).scalars().all()
    member_ids = [m.id for m in members]

    if not member_ids:
        return ExportData(
            sessions=list(sessions), members=[], att_map={},
            assign_map={}, team_ppt_map={}, ledger_by_session={},
        )

    # 3. 출석
    attendances = (await db.execute(
        select(Attendance).where(Attendance.member_id.in_(member_ids))
    )).scalars().all()
    att_map: dict[int, dict[int, Any]] = defaultdict(dict)
    for a in attendances:
        att_map[a.member_id][a.session_id] = a

    # 4. 과제 (개인)
    assignments = (await db.execute(
        select(Assignment).where(Assignment.member_id.in_(member_ids))
    )).scalars().all()
    assign_map: dict[int, dict[int, dict[str, Any]]] = defaultdict(lambda: defaultdict(dict))
    for a in assignments:
        assign_map[a.member_id][a.session_id][a.type] = a

    # 4.5 팀 PPT (member_id=NULL) → TeamMember 조인
    team_sessions = [s for s in sessions if s.type == "TEAM"]
    team_ppt_map: dict[int, dict[int, str]] = defaultdict(dict)  # session_id → member_id → status
    if team_sessions:
        team_session_ids = [s.id for s in team_sessions]
        team_ppts = (await db.execute(
            select(Assignment).where(
                Assignment.session_id.in_(team_session_ids),
                Assignment.type == "PPT_EMAIL",
                Assignment.member_id.is_(None),
            )
        )).scalars().all()
        for tp in team_ppts:
            if tp.team_id is None:
                continue
            tm_rows = (await db.execute(
                select(TeamMember.member_id).where(TeamMember.team_id == tp.team_id)
            )).all()
            for (mid,) in tm_rows:
                team_ppt_map[tp.session_id][mid] = tp.status

    # 5. 장부
    all_entries = (await db.execute(
        select(Ledger)
        .where(Ledger.member_id.in_(member_ids))
        .order_by(Ledger.created_at.asc())
    )).scalars().all()

    ledger_by_session: dict[int, dict[int, list]] = defaultdict(lambda: defaultdict(list))
    ledger_no_session: dict[int, list] = defaultdict(list)
    for e in all_entries:
        if e.session_id:
            ledger_by_session[e.member_id][e.session_id].append(e)
        else:
            ledger_no_session[e.member_id].append(e)

    return ExportData(
        sessions=list(sessions),
        members=list(members),
        att_map=dict(att_map),
        assign_map=dict(assign_map),
        team_ppt_map=dict(team_ppt_map),
        ledger_by_session=dict(ledger_by_session),
        ledger_no_session=dict(ledger_no_session),
    )


# ── 워크북 빌드 ──────────────────────────────────────────────────────

@dataclass
class WorkbookResult:
    wb: Workbook
    unmatched_merits: list[str]  # 엑셀 컬럼에 매칭 안 된 상점 description 목록


def build_workbook(data: ExportData, generation: int = 33) -> WorkbookResult:
    wb = Workbook()

    # 시트 순서대로 생성
    ws_penalty = wb.active
    ws_penalty.title = f"{generation}기 벌점"

    ws_deposit = wb.create_sheet(f"{generation}기 디파짓")
    ws_ppt = wb.create_sheet("PPT 제출")
    ws_att = wb.create_sheet("출석")
    ws_assign = wb.create_sheet("과제")
    ws_summary = wb.create_sheet("디파짓, 벌점 정리")

    sessions = data.sessions
    members = data.members

    unmatched = _build_penalty_sheet(ws_penalty, data, generation)
    _build_deposit_sheet(ws_deposit, data, generation)
    _build_ppt_sheet(ws_ppt, data, sessions, members)
    _build_attendance_sheet(ws_att, data, sessions, members)
    _build_assignment_sheet(ws_assign, data, sessions, members)
    _build_summary_sheet(ws_summary, data, sessions, members)

    return WorkbookResult(wb=wb, unmatched_merits=unmatched)


# ── 공통 헤더 작성 ───────────────────────────────────────────────────

def _write_common_headers(ws, sessions: list, title: str, extra_headers_r8: list[str] | None = None,
                          extra_headers_r10: list[str] | None = None):
    """3행 헤더 (날짜/주차/내용) + 타이틀 작성. extra_headers는 세션 컬럼 뒤에 추가할 요약 헤더."""
    # 타이틀 (Row 4, Col D)
    cell = ws.cell(row=4, column=NAME_COL, value=title)
    cell.font = TITLE_FONT

    # Row 8: 날짜
    ws.cell(row=HEADER_ROW, column=NAME_COL, value="날짜").font = HEADER_FONT
    for i, s in enumerate(sessions):
        col = DATA_START_COL + i
        date_str = s.date.strftime("%y.%m.%d") if s.date else ""
        c = ws.cell(row=HEADER_ROW, column=col, value=date_str)
        c.font = HEADER_FONT
        c.alignment = CENTER_ALIGN

    # Row 9: 주차
    ws.cell(row=WEEK_ROW, column=NAME_COL, value="주차").font = HEADER_FONT
    for i, s in enumerate(sessions):
        col = DATA_START_COL + i
        c = ws.cell(row=WEEK_ROW, column=col, value=s.week_num)
        c.font = HEADER_FONT
        c.alignment = CENTER_ALIGN

    # Row 10: 내용 (세션 타이틀)
    ws.cell(row=TITLE_ROW, column=NAME_COL, value="내용").font = HEADER_FONT
    for i, s in enumerate(sessions):
        col = DATA_START_COL + i
        c = ws.cell(row=TITLE_ROW, column=col, value=s.title or "")
        c.font = HEADER_FONT
        c.alignment = CENTER_ALIGN

    summary_start = DATA_START_COL + len(sessions)

    # extra headers on row 8
    if extra_headers_r8:
        for j, hdr in enumerate(extra_headers_r8):
            if hdr:
                c = ws.cell(row=HEADER_ROW, column=summary_start + j, value=hdr)
                c.font = HEADER_FONT
                c.alignment = WRAP_ALIGN

    # extra headers on row 10
    if extra_headers_r10:
        for j, hdr in enumerate(extra_headers_r10):
            if hdr:
                c = ws.cell(row=TITLE_ROW, column=summary_start + j, value=hdr)
                c.font = HEADER_FONT
                c.alignment = WRAP_ALIGN

    # 헤더 행 스타일
    for row_num in (HEADER_ROW, WEEK_ROW, TITLE_ROW):
        for col in range(NAME_COL, summary_start + len(extra_headers_r10 or extra_headers_r8 or [])):
            c = ws.cell(row=row_num, column=col)
            c.fill = HEADER_FILL
            c.border = THIN_BORDER

    return summary_start


def _write_member_column(ws, members: list):
    """C=번호, D=이름 작성"""
    for idx, m in enumerate(members):
        row = DATA_START_ROW + idx
        c_num = ws.cell(row=row, column=NUM_COL, value=idx + 1)
        c_num.alignment = CENTER_ALIGN
        c_num.border = THIN_BORDER
        c_name = ws.cell(row=row, column=NAME_COL, value=m.name)
        c_name.border = THIN_BORDER


def _apply_data_borders(ws, members: list, col_start: int, col_end: int):
    """데이터 영역에 테두리 적용"""
    for idx in range(len(members)):
        row = DATA_START_ROW + idx
        for col in range(col_start, col_end + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER


def _set_column_widths(ws, sessions: list, extra_count: int = 0):
    """컬럼 너비 설정"""
    ws.column_dimensions[get_column_letter(NUM_COL)].width = 5
    ws.column_dimensions[get_column_letter(NAME_COL)].width = 10
    for i in range(len(sessions)):
        ws.column_dimensions[get_column_letter(DATA_START_COL + i)].width = 14
    summary_start = DATA_START_COL + len(sessions)
    for j in range(extra_count):
        ws.column_dimensions[get_column_letter(summary_start + j)].width = 13


# ── Sheet 1: 벌점 ───────────────────────────────────────────────────

def _classify_merit(description: str) -> int | None:
    """상점 description을 카테고리 인덱스로 분류. 미매칭 시 None."""
    desc_lower = description
    for idx, (_, pattern) in enumerate(MERIT_CATEGORIES):
        # BP 매칭 시 "Listen Up/BP"와 구분
        if pattern == "BP":
            if "BP" in desc_lower and "Listen Up" not in desc_lower:
                return idx
        elif pattern in desc_lower:
            return idx
    return None


def _build_penalty_sheet(ws, data: ExportData, generation: int) -> list[str]:
    """벌점 시트 빌드. 미매칭 상점 description 목록 반환."""
    sessions = data.sessions
    members = data.members
    unmatched_descriptions: set[str] = set()

    # 상점 카테고리 헤더
    merit_headers = ["상점"] + [cat[0] for cat in MERIT_CATEGORIES]
    # 누적 컬럼: 상점 뒤에 빈 칸 두고 누적벌점, 누적상점, 상벌점총점
    gap = len(merit_headers)
    summary_r8 = [""] * gap + ["누적 벌점", "누적 상점", "상벌점\n총점"]
    summary_r10 = merit_headers + ["", "", ""]

    summary_start = _write_common_headers(
        ws, sessions, f"UnivPT {generation}기 벌점",
        extra_headers_r8=summary_r8,
        extra_headers_r10=summary_r10,
    )
    _write_member_column(ws, members)

    for idx, m in enumerate(members):
        row = DATA_START_ROW + idx
        member_ledger = data.ledger_by_session.get(m.id, {})

        # 주차별 벌점
        for si, s in enumerate(sessions):
            col = DATA_START_COL + si
            entries = member_ledger.get(s.id, [])
            penalty_score = sum(abs(e.score_delta) for e in entries if e.type == "FINE" and e.score_delta < 0)
            if penalty_score:
                c = ws.cell(row=row, column=col, value=penalty_score)
                c.alignment = CENTER_ALIGN
            ws.cell(row=row, column=col).border = THIN_BORDER

        # 상점 카테고리별 집계
        all_merits = []
        for sid_entries in member_ledger.values():
            all_merits.extend(e for e in sid_entries if e.type == "MERIT")
        # 세션 외 상점도 포함
        all_merits.extend(e for e in data.ledger_no_session.get(m.id, []) if e.type == "MERIT")

        merit_scores = [0] * len(MERIT_CATEGORIES)
        for e in all_merits:
            cat_idx = _classify_merit(e.description or "")
            if cat_idx is not None:
                merit_scores[cat_idx] += e.score_delta
            else:
                unmatched_descriptions.add(e.description or "(사유 없음)")

        # "상점" 라벨 컬럼은 비움, 카테고리별 점수 기입
        for ci, score in enumerate(merit_scores):
            col = summary_start + 1 + ci  # +1: "상점" 라벨 다음
            if score:
                c = ws.cell(row=row, column=col, value=score)
                c.alignment = CENTER_ALIGN
            ws.cell(row=row, column=col).border = THIN_BORDER

        # 누적 벌점, 누적 상점, 상벌점 총점
        cum_col = summary_start + gap
        penalty_val = abs(m.total_minus_score)
        ws.cell(row=row, column=cum_col, value=penalty_val if penalty_val else 0).alignment = CENTER_ALIGN
        ws.cell(row=row, column=cum_col).border = THIN_BORDER
        ws.cell(row=row, column=cum_col + 1, value=m.total_plus_score).alignment = CENTER_ALIGN
        ws.cell(row=row, column=cum_col + 1).border = THIN_BORDER
        net = m.net_score
        c = ws.cell(row=row, column=cum_col + 2, value=net if net else 0)
        c.alignment = CENTER_ALIGN
        c.border = THIN_BORDER

    total_extra = len(summary_r8)
    _set_column_widths(ws, sessions, total_extra)

    return sorted(unmatched_descriptions)


# ── Sheet 2: 디파짓 ──────────────────────────────────────────────────

def _build_deposit_sheet(ws, data: ExportData, generation: int):
    sessions = data.sessions
    members = data.members

    summary_r8 = ["총점"]
    summary_r10 = ["벌금", "부과된 디파짓", "충전한 금액", "내실 돈", "받으실 돈"]
    # row 8에는 "총점" 하나만, 나머지 4개는 row 10에만
    summary_r8_full = summary_r8 + [""] * (len(summary_r10) - 1)

    summary_start = _write_common_headers(
        ws, sessions, f"UnivPT {generation}기 디파짓",
        extra_headers_r8=summary_r8_full,
        extra_headers_r10=summary_r10,
    )
    _write_member_column(ws, members)

    for idx, m in enumerate(members):
        row = DATA_START_ROW + idx
        member_ledger = data.ledger_by_session.get(m.id, {})
        no_session_entries = data.ledger_no_session.get(m.id, [])

        total_fine_deposit = 0     # FINE 유형 디파짓 차감 합
        total_milestone = 0        # MILESTONE_FINE 합
        total_recharge = 0         # DEPOSIT_RECHARGE 합

        # 주차별 디파짓 차감
        for si, s in enumerate(sessions):
            col = DATA_START_COL + si
            entries = member_ledger.get(s.id, [])
            deposit_deduct = sum(abs(e.amount_krw) for e in entries if e.type == "FINE" and e.amount_krw < 0)
            if deposit_deduct:
                c = ws.cell(row=row, column=col, value=deposit_deduct)
                c.alignment = CENTER_ALIGN
                c.number_format = "#,##0"
            ws.cell(row=row, column=col).border = THIN_BORDER
            total_fine_deposit += deposit_deduct

        # 전체 세션 마일스톤 벌금
        for sid_entries in member_ledger.values():
            total_milestone += sum(abs(e.amount_krw) for e in sid_entries if e.type == "MILESTONE_FINE" and e.amount_krw < 0)
        # 세션 외 마일스톤
        total_milestone += sum(abs(e.amount_krw) for e in no_session_entries if e.type == "MILESTONE_FINE" and e.amount_krw < 0)

        # 충전
        for sid_entries in member_ledger.values():
            total_recharge += sum(e.amount_krw for e in sid_entries if e.type == "DEPOSIT_RECHARGE" and e.amount_krw > 0)
        total_recharge += sum(e.amount_krw for e in no_session_entries if e.type == "DEPOSIT_RECHARGE" and e.amount_krw > 0)

        total_deducted = total_fine_deposit + total_milestone  # 총 부과액

        # 요약 컬럼
        sc = summary_start
        # 벌금 (마일스톤)
        c = ws.cell(row=row, column=sc, value=total_milestone if total_milestone else None)
        c.alignment = CENTER_ALIGN
        c.number_format = "#,##0"
        c.border = THIN_BORDER
        # 부과된 디파짓
        c = ws.cell(row=row, column=sc + 1, value=total_deducted if total_deducted else 0)
        c.alignment = CENTER_ALIGN
        c.number_format = "#,##0"
        c.border = THIN_BORDER
        # 충전한 금액
        c = ws.cell(row=row, column=sc + 2, value=total_recharge if total_recharge else 0)
        c.alignment = CENTER_ALIGN
        c.number_format = "#,##0"
        c.border = THIN_BORDER
        # 내실 돈
        to_pay = max(total_deducted - total_recharge, 0)
        c = ws.cell(row=row, column=sc + 3, value=to_pay)
        c.alignment = CENTER_ALIGN
        c.number_format = "#,##0"
        c.border = THIN_BORDER
        # 받으실 돈
        to_receive = max(20000 - total_deducted + total_recharge, 0)
        c = ws.cell(row=row, column=sc + 4, value=to_receive)
        c.alignment = CENTER_ALIGN
        c.number_format = "#,##0"
        c.border = THIN_BORDER

    _set_column_widths(ws, sessions, len(summary_r10))


# ── Sheet 3: PPT 제출 ────────────────────────────────────────────────

def _build_ppt_sheet(ws, data: ExportData, sessions: list, members: list):
    _write_common_headers(ws, sessions, "PPT 제출")
    ws.cell(row=4, column=NAME_COL + 6, value="금요일 21:59:59 이후 제출 벌점 1점 / 미제출 벌점 2점").font = Font(size=9, color="666666")
    _write_member_column(ws, members)

    for idx, m in enumerate(members):
        row = DATA_START_ROW + idx
        for si, s in enumerate(sessions):
            col = DATA_START_COL + si
            # 개인 PPT_EMAIL
            assign = data.assign_map.get(m.id, {}).get(s.id, {}).get("PPT_EMAIL")
            status = None
            if assign:
                status = assign.status
            # 팀 PPT (개인 없을 때)
            if not status or status == "PENDING":
                team_status = data.team_ppt_map.get(s.id, {}).get(m.id)
                if team_status:
                    status = team_status

            text = PPT_STATUS_TEXT.get(status, "") if status else ""
            if text:
                c = ws.cell(row=row, column=col, value=text)
                c.alignment = CENTER_ALIGN
            ws.cell(row=row, column=col).border = THIN_BORDER

    _set_column_widths(ws, sessions)


# ── Sheet 4: 출석 ────────────────────────────────────────────────────

def _get_attendance_text(att) -> str:
    """Attendance 객체 → 한국어 텍스트"""
    if not att or att.status in ("PRESENT", "PENDING"):
        return ""
    key = (att.status, att.excuse_type if att.excuse_type else None)
    return ATTENDANCE_TEXT.get(key, att.status)


def _build_attendance_sheet(ws, data: ExportData, sessions: list, members: list):
    _write_common_headers(ws, sessions, "출석")
    _write_member_column(ws, members)

    for idx, m in enumerate(members):
        row = DATA_START_ROW + idx
        for si, s in enumerate(sessions):
            col = DATA_START_COL + si
            att = data.att_map.get(m.id, {}).get(s.id)
            text = _get_attendance_text(att)
            if text:
                c = ws.cell(row=row, column=col, value=text)
                c.alignment = WRAP_ALIGN
            ws.cell(row=row, column=col).border = THIN_BORDER

    _set_column_widths(ws, sessions)
    # 출석 컬럼은 줄바꿈 때문에 좀 더 넓게
    for i in range(len(sessions)):
        ws.column_dimensions[get_column_letter(DATA_START_COL + i)].width = 18


# ── Sheet 5: 과제 ────────────────────────────────────────────────────

def _build_assignment_sheet(ws, data: ExportData, sessions: list, members: list):
    _write_common_headers(ws, sessions, "과제")
    ws.cell(row=4, column=NAME_COL + 6, value="수요일 21:59:59 이후 제출 & 미제출 벌점 1점").font = Font(size=9, color="666666")
    _write_member_column(ws, members)

    for idx, m in enumerate(members):
        row = DATA_START_ROW + idx
        for si, s in enumerate(sessions):
            col = DATA_START_COL + si
            config = s.config or {}
            m_assigns = data.assign_map.get(m.id, {}).get(s.id, {})

            parts = []
            # 리뷰
            if config.get("has_review"):
                a = m_assigns.get("REVIEW")
                parts.append(f"리뷰{_status_mark(a)}")
            else:
                parts.append("리뷰-")
            # PPT (게시판)
            if config.get("has_ppt"):
                a = m_assigns.get("PPT")
                parts.append(f"PPT{_status_mark(a)}")
            else:
                parts.append("PPT-")
            # 과제
            if config.get("has_homework"):
                a = m_assigns.get("HOMEWORK")
                parts.append(f"과제{_status_mark(a)}")
            else:
                parts.append("과제-")

            text = "/".join(parts)
            # 전부 "-"면 빈 셀
            if all(p.endswith("-") for p in parts):
                text = ""

            if text:
                c = ws.cell(row=row, column=col, value=text)
                c.alignment = CENTER_ALIGN
                c.font = Font(size=9)
            ws.cell(row=row, column=col).border = THIN_BORDER

    _set_column_widths(ws, sessions)


def _status_mark(assignment) -> str:
    """Assignment → O/X/-"""
    if not assignment:
        return "-"
    if assignment.status == "PASS":
        return "O"
    if assignment.status == "MISSING":
        return "X"
    if assignment.status == "EXEMPT":
        return "-"
    # PENDING, LATE 등
    if assignment.status == "LATE":
        return "O"  # 지각 제출도 제출은 함
    return "-"


# ── Sheet 6: 디파짓, 벌점 정리 ────────────────────────────────────────

def _build_summary_sheet(ws, data: ExportData, sessions: list, members: list):
    # ── 상단: 디파짓 정리 ──
    _write_section_headers(ws, sessions, "디파짓 정리", start_row=4)
    _write_section_members(ws, members, start_row=4 + 7)  # row 11

    data_row_start = 4 + 7  # 11
    for idx, m in enumerate(members):
        row = data_row_start + idx
        for si, s in enumerate(sessions):
            col = DATA_START_COL + si
            att = data.att_map.get(m.id, {}).get(s.id)
            att_text = _get_attendance_text(att)

            # 해당 세션의 총 디파짓 차감
            entries = data.ledger_by_session.get(m.id, {}).get(s.id, [])
            total_deposit = sum(abs(e.amount_krw) for e in entries if e.type == "FINE" and e.amount_krw < 0)
            # PPT 벌금도 포함
            ppt_deposit = sum(abs(e.amount_krw) for e in entries if e.type == "FINE" and e.amount_krw < 0 and "PPT" in (e.description or ""))

            if att_text and total_deposit:
                # 출석 관련 금액
                att_deposit = ATTENDANCE_DEPOSIT.get(
                    (att.status, att.excuse_type if att.excuse_type else None), 0
                ) if att else 0
                cell_text = f"{att_text} {att_deposit:,}" if att_deposit else att_text
                # PPT 벌금이 별도로 있으면 추가
                other_deposit = total_deposit - att_deposit
                if other_deposit > 0:
                    cell_text += f"\nPPT {other_deposit:,}"
                c = ws.cell(row=row, column=col, value=cell_text)
                c.alignment = WRAP_ALIGN
            elif total_deposit:
                # 출석 이상 없이 PPT 등 과제 벌금만
                desc_parts = []
                for e in entries:
                    if e.type == "FINE" and e.amount_krw < 0:
                        desc_parts.append(f"{e.description} {abs(e.amount_krw):,}")
                c = ws.cell(row=row, column=col, value="\n".join(desc_parts))
                c.alignment = WRAP_ALIGN
            ws.cell(row=row, column=col).border = THIN_BORDER

    # ── 하단: 벌점 정리 ──
    gap_row = data_row_start + len(members) + 6  # 몇 줄 간격
    _write_section_headers(ws, sessions, "벌점 정리", start_row=gap_row)
    penalty_data_start = gap_row + 7

    for idx, m in enumerate(members):
        row = penalty_data_start + idx
        # 번호, 이름
        ws.cell(row=row, column=NUM_COL, value=idx + 1).alignment = CENTER_ALIGN
        ws.cell(row=row, column=NUM_COL).border = THIN_BORDER
        ws.cell(row=row, column=NAME_COL, value=m.name).border = THIN_BORDER

        for si, s in enumerate(sessions):
            col = DATA_START_COL + si
            entries = data.ledger_by_session.get(m.id, {}).get(s.id, [])
            penalty_entries = [e for e in entries if e.type == "FINE" and e.score_delta < 0]

            if penalty_entries:
                parts = []
                for e in penalty_entries:
                    parts.append(f"{e.description} 벌점{abs(e.score_delta)}")
                c = ws.cell(row=row, column=col, value="\n".join(parts))
                c.alignment = WRAP_ALIGN
                c.font = Font(size=9)
            ws.cell(row=row, column=col).border = THIN_BORDER

    # 컬럼 너비
    ws.column_dimensions[get_column_letter(NUM_COL)].width = 5
    ws.column_dimensions[get_column_letter(NAME_COL)].width = 10
    for i in range(len(sessions)):
        ws.column_dimensions[get_column_letter(DATA_START_COL + i)].width = 20


def _write_section_headers(ws, sessions: list, title: str, start_row: int):
    """디파짓/벌점 정리 시트의 섹션 헤더"""
    ws.cell(row=start_row, column=NAME_COL, value=title).font = TITLE_FONT

    hr = start_row + 4  # 날짜 행
    wr = start_row + 5  # 주차 행
    tr = start_row + 6  # 내용 행

    ws.cell(row=hr, column=NAME_COL, value="날짜").font = HEADER_FONT
    ws.cell(row=wr, column=NAME_COL, value="주차").font = HEADER_FONT
    ws.cell(row=tr, column=NAME_COL, value="내용").font = HEADER_FONT

    for i, s in enumerate(sessions):
        col = DATA_START_COL + i
        date_str = s.date.strftime("%y.%m.%d") if s.date else ""
        ws.cell(row=hr, column=col, value=date_str).font = HEADER_FONT
        ws.cell(row=hr, column=col).alignment = CENTER_ALIGN
        ws.cell(row=hr, column=col).fill = HEADER_FILL
        ws.cell(row=hr, column=col).border = THIN_BORDER

        ws.cell(row=wr, column=col, value=s.week_num).font = HEADER_FONT
        ws.cell(row=wr, column=col).alignment = CENTER_ALIGN
        ws.cell(row=wr, column=col).fill = HEADER_FILL
        ws.cell(row=wr, column=col).border = THIN_BORDER

        ws.cell(row=tr, column=col, value=s.title or "").font = HEADER_FONT
        ws.cell(row=tr, column=col).alignment = CENTER_ALIGN
        ws.cell(row=tr, column=col).fill = HEADER_FILL
        ws.cell(row=tr, column=col).border = THIN_BORDER

    # 이름/날짜/주차/내용 컬럼 헤더 스타일
    for r in (hr, wr, tr):
        ws.cell(row=r, column=NAME_COL).fill = HEADER_FILL
        ws.cell(row=r, column=NAME_COL).border = THIN_BORDER


def _write_section_members(ws, members: list, start_row: int):
    """디파짓/벌점 정리 시트의 멤버 번호+이름"""
    for idx, m in enumerate(members):
        row = start_row + idx
        ws.cell(row=row, column=NUM_COL, value=idx + 1).alignment = CENTER_ALIGN
        ws.cell(row=row, column=NUM_COL).border = THIN_BORDER
        ws.cell(row=row, column=NAME_COL, value=m.name).border = THIN_BORDER


# ── 엔트리포인트 ─────────────────────────────────────────────────────

@dataclass
class ExcelExportResult:
    stream: io.BytesIO
    unmatched_merits: list[str]


def check_unmatched_merits(data: ExportData) -> list[dict]:
    """엑셀 컬럼에 매칭 안 되는 상점 항목 반환 (id, member_name, description)"""
    unmatched: list[dict] = []
    member_name_map = {m.id: m.name for m in data.members}
    for m in data.members:
        all_merits = []
        for sid_entries in data.ledger_by_session.get(m.id, {}).values():
            all_merits.extend(e for e in sid_entries if e.type == "MERIT")
        all_merits.extend(e for e in data.ledger_no_session.get(m.id, []) if e.type == "MERIT")
        for e in all_merits:
            if _classify_merit(e.description or "") is None:
                unmatched.append({
                    "id": e.id,
                    "member_name": member_name_map.get(e.member_id, ""),
                    "description": e.description or "(사유 없음)",
                    "score_delta": e.score_delta,
                })
    return unmatched


async def generate_excel_bytes(db: AsyncSession, generation: int = 33) -> ExcelExportResult:
    """DB에서 데이터 수집 → Excel 워크북 생성 → BytesIO + 미매칭 상점 목록 반환"""
    data = await gather_export_data(db)
    result = build_workbook(data, generation=generation)
    stream = io.BytesIO()
    result.wb.save(stream)
    stream.seek(0)
    return ExcelExportResult(stream=stream, unmatched_merits=result.unmatched_merits)
