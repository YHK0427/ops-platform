"""심사 결과 Excel 내보내기.

시트 5개:
  1. 결과      — 순위 / 팀 / 심사위원·참관위원 점수 / 총점 / 기준별 평균
  2. 심사위원별 — 행=제출자, 열=팀 (관대·엄격 편차 확인용 매트릭스)
  3. 상세점수  — 제출자 × 팀 × 기준 원점수 (raw dump, 재계산·검증용)
  4. 피드백    — 팀별 기준 코멘트 + 총평
  5. 제출현황  — 명단 대비 제출 여부 + 미매칭 제출자
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    ScoringComment, ScoringDeduction, ScoringParticipant, ScoringRank, ScoringRound, ScoringScore,
)
from app.services.scoring_engine import (
    AreaLite, CriterionLite, ParticipantLite, RankLite, ScoreLite, compute_results,
)

HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
TITLE_FONT = Font(bold=True, size=13)
WINNER_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
DQ_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # 실격
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

ROLE_KR = {"JUDGE": "심사위원", "OBSERVER": "청중"}


def _style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def _autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


async def generate_scoring_excel(db: AsyncSession, rnd: ScoringRound) -> io.BytesIO:
    parts = (await db.execute(
        select(ScoringParticipant).where(ScoringParticipant.round_id == rnd.id)
        .order_by(ScoringParticipant.created_at)
    )).scalars().all()
    submitted = [p for p in parts if p.submitted_at is not None]
    pids = [p.id for p in submitted]

    scores = (await db.execute(
        select(ScoringScore).where(ScoringScore.participant_id.in_(pids))
    )).scalars().all() if pids else []
    ranks = (await db.execute(
        select(ScoringRank).where(ScoringRank.participant_id.in_(pids))
    )).scalars().all() if pids else []
    comments = (await db.execute(
        select(ScoringComment).where(ScoringComment.participant_id.in_(pids))
    )).scalars().all() if pids else []

    targets = sorted(rnd.targets, key=lambda t: t.order_num)
    criteria = sorted(rnd.criteria, key=lambda c: c.order_num)
    areas = sorted(rnd.areas, key=lambda a: a.order_num)

    # 팀별 감점·실격
    dedns = (await db.execute(
        select(ScoringDeduction).where(ScoringDeduction.round_id == rnd.id)
    )).scalars().all()
    rule_label = {r.id: r.label for r in rnd.deduction_rules}
    deduction_by_target: dict[int, float] = {}
    disqualified: set[int] = set()
    for d in dedns:
        if d.disqualified:
            disqualified.add(d.target_id)
        deduction_by_target[d.target_id] = deduction_by_target.get(d.target_id, 0.0) + float(d.points)

    computed = compute_results(
        judge_weight=float(rnd.judge_weight),
        observer_weight=float(rnd.observer_weight),
        observer_mode=rnd.observer_mode,
        rank_points=list(rnd.rank_points or []),
        criteria=[CriterionLite(id=c.id, max_score=float(c.max_score), area_id=c.area_id) for c in criteria],
        areas=[
            AreaLite(id=a.id, max_score=float(a.max_score),
                     criterion_ids=tuple(c.id for c in criteria if c.area_id == a.id))
            for a in areas
        ],
        target_ids=[t.id for t in targets],
        participants=[ParticipantLite(id=p.id, role=p.role, name=p.entered_name) for p in submitted],
        scores=[ScoreLite(participant_id=s.participant_id, target_id=s.target_id,
                          criterion_id=s.criterion_id, area_id=s.area_id, score=float(s.score))
                for s in scores],
        ranks=[RankLite(participant_id=r.participant_id, target_id=r.target_id, rank=r.rank) for r in ranks],
        deductions=deduction_by_target,
        disqualified=disqualified,
    )

    # 결과/심사위원별 시트의 점수 컬럼 = 영역별 평균 + 미분류 기준별 평균
    ungrouped = [c for c in criteria if c.area_id is None]
    # (label, getter) 목록 — 영역은 area_avg, 미분류는 criterion_avg
    score_cols = [(a.label, ("area", a.id)) for a in areas] + \
                 [(c.label, ("crit", c.id)) for c in ungrouped]

    def col_val(r, key) -> float:
        kind, kid = key
        return (r.area_avg.get(kid, 0) if kind == "area" else r.criterion_avg.get(kid, 0))

    # 표시명을 정했으면 그걸 쓴다 (평가 폼·결과와 동일하게 보이도록)
    def disp(t) -> str:
        return (t.display_name or "").strip() or t.name

    tname = {t.id: disp(t) for t in targets}
    cname = {c.id: c.label for c in criteria}
    pname = {p.id: p.entered_name for p in submitted}
    prole = {p.id: p.role for p in submitted}

    wb = Workbook()

    # ── 1. 결과 ──
    ws = wb.active
    ws.title = "결과"
    ws["A1"] = rnd.name
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        f"심사위원 {float(rnd.judge_weight):g}점 · "
        f"참관위원 {float(rnd.observer_weight):g}점"
        f"({'등수 선택' if rnd.observer_mode == 'RANK' else '기준 채점'})"
    )
    ws["A3"] = (
        f"제출: 심사위원 {len([p for p in submitted if p.role == 'JUDGE'])}명 · "
        f"참관위원 {len([p for p in submitted if p.role == 'OBSERVER'])}명"
    )

    has_ded = bool(rnd.deduction_rules)
    head = ["순위", "팀", "심사위원", "참관위원"]
    if has_ded:
        head += ["감점전", "감점", "최종"]
    else:
        head += ["총점"]
    head += [label for label, _ in score_cols]
    ws.append([])
    ws.append(head)
    hrow = ws.max_row
    _style_header(ws, hrow, len(head))

    for r in computed:
        rank_cell = "실격" if r.disqualified else r.rank
        row = [rank_cell, tname.get(r.target_id, "?"), r.judge_points, r.observer_points]
        if has_ded:
            row += [r.pre_deduction, r.deduction, r.total]
        else:
            row += [r.total]
        row += [col_val(r, key) for _, key in score_cols]
        ws.append(row)
        for c in range(1, len(head) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.border = BORDER
            if c != 2:
                cell.alignment = CENTER
            if r.disqualified:
                cell.fill = DQ_FILL
            elif r.rank == 1:
                cell.fill = WINNER_FILL
    _autosize(ws, [6, 22, 11, 11] + ([9, 8, 9] if has_ded else [9]) + [12] * len(score_cols))

    # ── 2. 심사위원별 (매트릭스) ──
    ws2 = wb.create_sheet("심사위원별")
    ws2["A1"] = "제출자별 부여 점수 (기준 원점수 합계) — 빈칸은 채점하지 않은 팀"
    ws2["A1"].font = TITLE_FONT
    # 영역(area)은 세부항목이 있어도 없어도(통째만) 배점을 한 번만 센다 — criteria만 더하면
    # 세부항목 없는(통째 전용) 영역의 배점이 통째로 빠진다.
    max_total = sum(float(a.max_score) for a in areas) + sum(float(c.max_score) for c in ungrouped)
    ws2["A2"] = f"기준 만점 합계: {max_total:g}점"

    head2 = ["제출자", "역할"] + [disp(t) for t in targets]
    ws2.append([])
    ws2.append(head2)
    _style_header(ws2, ws2.max_row, len(head2))

    totals_by: dict[int, dict[int, float]] = {}
    for s in scores:
        totals_by.setdefault(s.participant_id, {})
        totals_by[s.participant_id][s.target_id] = (
            totals_by[s.participant_id].get(s.target_id, 0.0) + float(s.score)
        )

    for p in submitted:
        if p.role == "OBSERVER" and rnd.observer_mode == "RANK":
            continue  # 등수 모드 참관위원은 점수를 매기지 않는다
        row = [p.entered_name, ROLE_KR.get(p.role, p.role)]
        for t in targets:
            v = totals_by.get(p.id, {}).get(t.id)
            row.append(v if v is not None else "")
        ws2.append(row)
        for c in range(1, len(head2) + 1):
            cell = ws2.cell(row=ws2.max_row, column=c)
            cell.border = BORDER
            if c > 2:
                cell.alignment = CENTER
    _autosize(ws2, [16, 11] + [14] * len(targets))

    # 등수 모드 참관위원의 투표는 별도 블록으로
    if rnd.observer_mode == "RANK":
        ws2.append([])
        ws2.append(["참관위원 등수 투표"])
        ws2.cell(row=ws2.max_row, column=1).font = TITLE_FONT
        head2b = ["제출자", "그룹"] + [disp(t) for t in targets]
        ws2.append(head2b)
        _style_header(ws2, ws2.max_row, len(head2b))

        rank_by: dict[int, dict[int, int]] = {}
        for rk in ranks:
            rank_by.setdefault(rk.participant_id, {})[rk.target_id] = rk.rank
        # 그룹별로 묶어서 읽기 좋게 (운영진 / 기수 / 청중 …)
        for p in sorted(
            [x for x in submitted if x.role == "OBSERVER"],
            key=lambda x: (x.group_label or "￿", x.entered_name),
        ):
            row = [p.entered_name, p.group_label or "미분류"]
            for t in targets:
                rk = rank_by.get(p.id, {}).get(t.id)
                row.append(f"{rk}위" if rk else "")
            ws2.append(row)
            for c in range(1, len(head2b) + 1):
                cell = ws2.cell(row=ws2.max_row, column=c)
                cell.border = BORDER
                if c > 2:
                    cell.alignment = CENTER

    # ── 3. 상세점수 (raw dump) — 세부항목/미분류 + 영역 통째 ──
    ws3 = wb.create_sheet("상세점수")
    head3 = ["제출자", "역할", "팀", "심사기준", "배점", "점수"]
    ws3.append(head3)
    _style_header(ws3, 1, len(head3))
    maxes = {c.id: float(c.max_score) for c in criteria}
    amax = {a.id: float(a.max_score) for a in areas}
    alabel = {a.id: f"{a.label} (통째)" for a in areas}

    def _item_label(s) -> str:
        if s.criterion_id is not None:
            return cname.get(s.criterion_id, "?")
        return alabel.get(s.area_id, "?")

    def _item_max(s) -> float:
        return maxes.get(s.criterion_id, 0) if s.criterion_id is not None else amax.get(s.area_id, 0)

    for s in sorted(scores, key=lambda x: (x.participant_id, x.target_id,
                                           x.criterion_id or 0, x.area_id or 0)):
        ws3.append([
            pname.get(s.participant_id, "?"),
            ROLE_KR.get(prole.get(s.participant_id, ""), ""),
            tname.get(s.target_id, "?"),
            _item_label(s),
            _item_max(s),
            float(s.score),
        ])
        for c in range(1, len(head3) + 1):
            ws3.cell(row=ws3.max_row, column=c).border = BORDER
    _autosize(ws3, [16, 11, 20, 22, 8, 8])

    # ── 4. 피드백 ──
    ws4 = wb.create_sheet("피드백")
    head4 = ["팀", "구분", "제출자", "역할", "내용"]
    ws4.append(head4)
    _style_header(ws4, 1, len(head4))
    # 팀 순위 순으로 정렬해 읽기 좋게
    order = {r.target_id: r.rank for r in computed}
    for c in sorted(comments, key=lambda x: (order.get(x.target_id, 999), x.criterion_id or 0)):
        ws4.append([
            tname.get(c.target_id, "?"),
            cname.get(c.criterion_id, "총평") if c.criterion_id else "총평",
            pname.get(c.participant_id, "?"),
            ROLE_KR.get(prole.get(c.participant_id, ""), ""),
            c.body,
        ])
        for col in range(1, len(head4) + 1):
            cell = ws4.cell(row=ws4.max_row, column=col)
            cell.border = BORDER
            if col == 5:
                cell.alignment = WRAP
    _autosize(ws4, [20, 18, 16, 11, 80])

    # ── 5. 제출현황 ──
    # 대리 입력 여부·명단 미매칭 같은 운영 내부 사정은 싣지 않는다 (결과물이니까).
    ws5 = wb.create_sheet("제출현황")
    head5 = ["명단 이름", "역할", "그룹", "비고", "제출", "제출시각"]
    ws5.append(head5)
    _style_header(ws5, 1, len(head5))

    by_roster: dict[int, ScoringParticipant] = {
        p.matched_roster_id: p for p in parts
        if p.matched_roster_id is not None and p.submitted_at is not None
    }
    for entry in rnd.roster:
        p = by_roster.get(entry.id)
        ws5.append([
            entry.name,
            ROLE_KR.get(entry.role, "무관"),
            (p.group_label if p else "") or "",
            entry.note or "",
            "제출" if p else "미제출",
            p.submitted_at.strftime("%Y-%m-%d %H:%M") if p and p.submitted_at else "",
        ])
        for c in range(1, len(head5) + 1):
            ws5.cell(row=ws5.max_row, column=c).border = BORDER
    _autosize(ws5, [16, 11, 12, 18, 9, 18])

    # ── 6. 감점 (규정이 있을 때만) ──
    if rnd.deduction_rules:
        ws6 = wb.create_sheet("감점")
        rules = sorted(rnd.deduction_rules, key=lambda r: r.order_num)
        head6 = ["팀"] + [r.label for r in rules] + ["감점 합계", "실격"]
        ws6.append(head6)
        _style_header(ws6, 1, len(head6))

        # (target_id, rule_id) → (points, disqualified)
        cell_map: dict[tuple[int, int], tuple[float, bool]] = {
            (d.target_id, d.rule_id): (float(d.points), d.disqualified) for d in dedns
        }
        for t in targets:
            total_ded = deduction_by_target.get(t.id, 0.0)
            row = [tname.get(t.id, "?")]
            for r in rules:
                pts, dq = cell_map.get((t.id, r.id), (0.0, False))
                row.append("실격" if dq else (pts if pts else ""))
            row.append(total_ded)
            row.append("실격" if t.id in disqualified else "")
            ws6.append(row)
            for c in range(1, len(head6) + 1):
                cell = ws6.cell(row=ws6.max_row, column=c)
                cell.border = BORDER
                cell.alignment = CENTER
                if t.id in disqualified:
                    cell.fill = DQ_FILL
        _autosize(ws6, [20] + [14] * len(rules) + [11, 8])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
